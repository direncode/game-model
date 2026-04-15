"""Excel connector — read .xlsx/.xls files."""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from ..base import BaseDatasetAdapter, DatasetMeta, EntityTypeConfig

logger = logging.getLogger(__name__)


def _singularize(name: str) -> str:
    if not name:
        return name
    lower = name.lower()
    if lower.endswith("ies") and len(lower) > 3:
        return name[:-3] + "y"
    if lower.endswith("s") and not lower.endswith("ss"):
        return name[:-1]
    return name


class ExcelConnector(BaseDatasetAdapter):
    """Connect to Excel files. Each sheet becomes an entity type.

    Requires ``openpyxl``. Install with: ``pip install openpyxl``

    Usage::

        connector = ExcelConnector("/path/to/data.xlsx")
        connector = ExcelConnector("/path/to/data.xlsx", sheets=["Sheet1", "Data"])
    """

    def __init__(
        self,
        source: str,
        sheets: list[str] | None = None,
        name_column: str | None = None,
        encoding: str = "utf-8",
    ) -> None:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel support. "
                "Install with: pip install openpyxl"
            )

        self._path = Path(source)
        if not self._path.exists():
            raise FileNotFoundError(f"Excel file not found: {source}")

        self._openpyxl = openpyxl
        self._name_column = name_column
        self._source_name = self._path.stem
        self._requested_sheets = sheets

        # Parse workbook
        self._sheet_data: dict[str, list[dict]] = {}
        self._sheet_headers: dict[str, list[str]] = {}
        self._parse()

    def _parse(self) -> None:
        wb = self._openpyxl.load_workbook(str(self._path), read_only=True, data_only=True)
        sheet_names = self._requested_sheets or wb.sheetnames

        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                logger.warning("Sheet '%s' not found in %s, skipping", sheet_name, self._path)
                continue

            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)

            # First row is headers
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue

            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
            self._sheet_headers[sheet_name] = headers

            records: list[dict] = []
            for row in rows_iter:
                record: dict[str, Any] = {}
                for i, value in enumerate(row):
                    if i < len(headers) and value is not None:
                        record[headers[i]] = value
                if record:
                    records.append(record)

            self._sheet_data[sheet_name] = records
            logger.info("Sheet '%s': %d rows, %d columns", sheet_name, len(records), len(headers))

        wb.close()

    def _pick_name_column(self, headers: list[str]) -> str:
        if self._name_column:
            return self._name_column
        priority = ["name", "title", "label", "key", "id"]
        lower_headers = {h.lower(): h for h in headers}
        for candidate in priority:
            if candidate in lower_headers:
                return lower_headers[candidate]
        return headers[0] if headers else ""

    # ------------------------------------------------------------------ #
    #  BaseDatasetAdapter interface
    # ------------------------------------------------------------------ #

    def get_meta(self) -> DatasetMeta:
        total_rows = sum(len(rows) for rows in self._sheet_data.values())
        all_columns: set[str] = set()
        for headers in self._sheet_headers.values():
            all_columns.update(headers)

        entity_types: list[EntityTypeConfig] = []
        colors = ["#00d4ff", "#ff6b6b", "#51cf66", "#ffd43b", "#845ef7", "#ff922b"]
        for i, (sheet_name, headers) in enumerate(self._sheet_headers.items()):
            name_col = self._pick_name_column(headers)
            secondary = ""
            for h in headers:
                if h != name_col:
                    secondary = h
                    break
            entity_types.append(
                EntityTypeConfig(
                    name=_singularize(sheet_name),
                    color=colors[i % len(colors)],
                    primary_field=name_col,
                    secondary_field=secondary,
                    detail_fields=headers[:8],
                )
            )

        return DatasetMeta(
            dataset_id=f"excel:{self._source_name}",
            display_name=f"Excel: {self._source_name}",
            description=f"{total_rows} rows across {len(self._sheet_data)} sheet(s)",
            entity_types=entity_types,
            lookup_field="name",
            lookup_label="Record",
        )

    def fetch_entities(self, limit: int = 10_000) -> list[dict]:
        entities: list[dict] = []
        remaining = limit

        for sheet_name, records in self._sheet_data.items():
            if remaining <= 0:
                break

            headers = self._sheet_headers.get(sheet_name, [])
            name_col = self._pick_name_column(headers)
            entity_type = _singularize(sheet_name)

            for idx, record in enumerate(records[:remaining]):
                if name_col and name_col in record:
                    name = str(record[name_col])
                else:
                    name = f"{entity_type}:{idx}"

                # Clean attributes
                attributes: dict[str, Any] = {}
                for k, v in record.items():
                    if v is not None:
                        # Convert datetime objects to ISO strings
                        if hasattr(v, "isoformat"):
                            attributes[k] = v.isoformat()
                        else:
                            attributes[k] = v

                entities.append({
                    "name": name,
                    "type": entity_type,
                    "attributes": attributes,
                })
                remaining -= 1

        logger.info("Produced %d entities from Excel", len(entities))
        return entities

    def fetch_edges(self, entities: list[dict]) -> list[dict]:
        """Infer edges from _id columns across sheets."""
        entity_index: dict[tuple[str, str], str] = {}
        for ent in entities:
            ename = ent["name"]
            etype = ent["type"]
            entity_index[(etype, ename)] = ename
            for k, v in ent.get("attributes", {}).items():
                if isinstance(v, (str, int, float)):
                    entity_index[(etype, str(v))] = ename

        all_headers: set[str] = set()
        for headers in self._sheet_headers.values():
            all_headers.update(headers)

        id_columns = [h for h in all_headers if h.endswith("_id")]
        edges: list[dict] = []

        for ent in entities:
            attrs = ent.get("attributes", {})
            for col in id_columns:
                ref_value = attrs.get(col)
                if ref_value is None:
                    continue
                ref_type = _singularize(col[:-3])
                target_key = (ref_type, str(ref_value))
                target_name = entity_index.get(target_key)
                if target_name and target_name != ent["name"]:
                    edges.append({
                        "source": ent["name"],
                        "target": target_name,
                        "type": f"{ent['type']}_to_{ref_type}",
                        "weight": 1.0,
                    })

        logger.info("Inferred %d edges from Excel _id columns", len(edges))
        return edges

    def get_anomaly_tags(
        self,
        record: dict,
        scores: dict,
        cluster_size: int,
        flips: int,
    ) -> list[dict]:
        from .csv_connector import _generic_anomaly_tags
        return _generic_anomaly_tags(record, scores, cluster_size, flips)
